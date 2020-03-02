from openpyxl import load_workbook

class ExcelReader:

	def readData(filePath,sheetName):
		wb=load_workbook(filename=filePath,read_only=True)
		ws=wb[sheetName]
		maxRow=ws.max_row
		maxColumn=ws.max_column
		print("Rows in sheet "+str(maxRow))
		print("Columns in sheet "+str(maxColumn))
		data=[]
		for i in range(maxRow-1):
			tempRowArray=[]
			for j in range(maxColumn):
				tempRowArray.append(ws.cell(i+2,j+1).value)
			data.append(tempRowArray)
		return data;